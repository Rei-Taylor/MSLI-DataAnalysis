import gspread
from oauth2client.client import OAuth2WebServerFlow
from oauth2client.tools import run_flow
from oauth2client.file import Storage
from tabulate import tabulate
from pyfiglet import Figlet
import re
from datetime import datetime
from openpyxl import Workbook

def print_logo():
    f = Figlet(font='slant')
    print(f.renderText('Rei-Taylor'))

# Function to print credit line
def print_credit():
    print("\nPowered by Rei-Taylor")

# Function to get client ID and client secret from user input
def get_credentials():
    client_id = input("Enter your Client ID: ")
    client_secret = input("Enter your Client Secret: ")
    return client_id, client_secret

# Define the scope of the Google Sheets API
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

# Get client ID and client secret
CLIENT_ID, CLIENT_SECRET = get_credentials()

# Create a flow object using the client ID, client secret, and the specified scope
flow = OAuth2WebServerFlow(client_id=CLIENT_ID, client_secret=CLIENT_SECRET, scope=scope, redirect_uri='urn:ietf:wg:oauth:2.0:oob')

# Run the OAuth2 authorization flow and get credentials
storage = Storage('credentials.dat')  # This will store the credentials in a local file
credentials = run_flow(flow, storage)

# Use the credentials to authorize gspread
gc = gspread.authorize(credentials)

def add_row(spreadsheet, sheet_name, values):
    worksheet = spreadsheet.worksheet(sheet_name)
    worksheet.append_row(values)
    print("Row added successfully to sheet '{}'.".format(sheet_name))

def delete_row(spreadsheet, sheet_name, row_number):
    worksheet = spreadsheet.worksheet(sheet_name)
    worksheet.delete_row(row_number)
    print("Row {} deleted successfully from sheet '{}'.".format(row_number, sheet_name))

def get_sheet_row_count(spreadsheet, sheet_name):
    worksheet = spreadsheet.worksheet(sheet_name)
    return worksheet.row_count

def get_sheet_values(spreadsheet, sheet_name):
    worksheet = spreadsheet.worksheet(sheet_name)
    return worksheet.get_all_values()

def display_table(values):
    headers = values[0]
    data = values[1:]
    numbered_data = [[i+1] + row for i, row in enumerate(data)]  # Add row numbers to each row
    print(tabulate(numbered_data, headers=['Row'] + headers, tablefmt="pretty"))

def analyze_data(spreadsheet, sheet_name):
    worksheet = spreadsheet.worksheet(sheet_name)
    column_names = worksheet.row_values(1)  # Get column names from the first row
    print("Columns in sheet '{}': {}".format(sheet_name, column_names))
    column_to_search = input("Enter the name of the column to search: ")
    if column_to_search not in column_names:
        print("Invalid column name. Please try again.")
        return

    value_to_find = input("Enter the value to find in column '{}': ".format(column_to_search))
    if not value_to_find:
        print("Invalid value. Please try again.")
        return

    total_kg_column = input("Enter the name of the 'Total kg' column: ")
    if total_kg_column not in column_names:
        print("Invalid 'Total kg' column name. Please try again.")
        return

    total_mc_column = input("Enter the name of the 'Total mc' column: ")
    if total_mc_column not in column_names:
        print("Invalid 'Total mc' column name. Please try again.")
        return

    use_date_filter = input("Would you like to use date filtering? (yes/no): ").lower()
    if use_date_filter == 'yes':
        date_column = input("Enter the name of the date column: ")
        if date_column not in column_names:
            print("Invalid date column name. Please try again.")
            return

        date_before = input("Enter the date before (MM/DD/YY): ")
        date_after = input("Enter the date after (MM/DD/YY): ")

        try:
            date_before = datetime.strptime(date_before, "%m/%d/%y")
            date_after = datetime.strptime(date_after, "%m/%d/%y")
        except ValueError:
            print("Invalid date format. Please enter the date in the format MM/DD/YY.")
            return

    data = get_sheet_values(spreadsheet, sheet_name)
    headers = data[0]
    column_index = [i for i, header in enumerate(headers) if header.lower() == column_to_search.lower()]
    if not column_index:
        print("Column '{}' not found in the sheet.".format(column_to_search))
        return
    
    column_index = column_index[0]
    relevant_rows = [row for row in data[1:] if row[column_index].lower() == value_to_find.lower()]

    if use_date_filter == 'yes':
        relevant_rows_within_dates = [
            row for row in relevant_rows if date_before <= datetime.strptime(row[column_names.index(date_column)], "%m/%d/%y") <= date_after
        ]

        if relevant_rows_within_dates:
            print("Number of rows with '{}' in column '{}' and within the specified dates: {}".format(
                value_to_find, column_to_search, len(relevant_rows_within_dates)))

            total_kg_index = headers.index(total_kg_column)
            total_mc_index = headers.index(total_mc_column)

            total_kg = sum(float(re.findall(r'\d+\.\d+|\d+', row[total_kg_index])[0]) for row in relevant_rows_within_dates if re.findall(r'\d+\.\d+|\d+', row[total_kg_index]))
            total_mc = sum(float(re.findall(r'\d+\.\d+|\d+', row[total_mc_index])[0]) for row in relevant_rows_within_dates if re.findall(r'\d+\.\d+|\d+', row[total_mc_index]))

            print("Total kg for '{}' in column '{}' within specified dates: {}".format(value_to_find, column_to_search, total_kg))
            print("Total mc for '{}' in column '{}' within specified dates: {}".format(value_to_find, column_to_search, total_mc))
        else:
            print("No rows found with '{}' in column '{}' within the specified dates.".format(value_to_find, column_to_search))
    else:
        if relevant_rows:
            print("Number of rows with '{}' in column '{}': {}".format(value_to_find, column_to_search, len(relevant_rows)))

            total_kg_index = headers.index(total_kg_column)
            total_mc_index = headers.index(total_mc_column)

            total_kg = sum(float(re.findall(r'\d+\.\d+|\d+', row[total_kg_index])[0]) for row in relevant_rows if re.findall(r'\d+\.\d+|\d+', row[total_kg_index]))
            total_mc = sum(float(re.findall(r'\d+\.\d+|\d+', row[total_mc_index])[0]) for row in relevant_rows if re.findall(r'\d+\.\d+|\d+', row[total_mc_index]))

            print("Total kg for '{}' in column '{}': {}".format(value_to_find, column_to_search, total_kg))
            print("Total mc for '{}' in column '{}': {}".format(value_to_find, column_to_search, total_mc))
        else:
            print("No rows found with '{}' in column '{}'.".format(value_to_find, column_to_search))

def export_to_xlsx(spreadsheet, sheet_name):
    data = get_sheet_values(spreadsheet, sheet_name)
    headers = data[0]
    rows = data[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook
    file_name = f"{sheet_name}.xlsx"
    wb.save(filename=file_name)
    print(f"The sheet '{sheet_name}' has been exported to '{file_name}'.")

# Main function
def main():
    print_logo()
    # Open the spreadsheet
    spreadsheet_name = input("Enter the name of the spreadsheet: ")
    spreadsheet = gc.open(spreadsheet_name)

    while True:
        print("\nOptions:")
        print("1. Add a row")
        print("2. Delete a row")
        print("3. Check the number of rows in a sheet")
        print("4. View sheet contents as a table")
        print("5. Analyze data")
        print("6. Export sheet to XLSX file")
        print("7. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            sheet_name = input("Enter the name of the sheet: ")
            values = input("Enter values for each column (comma-separated): ").split(',')
            add_row(spreadsheet, sheet_name, values)
        elif choice == '2':
            sheet_name = input("Enter the name of the sheet: ")
            row_number = int(input("Enter the row number to delete: "))
            delete_row(spreadsheet, sheet_name, row_number)
        elif choice == '3':
            sheet_name = input("Enter the name of the sheet: ")
            print("Number of rows in sheet '{}': {}".format(sheet_name, get_sheet_row_count(spreadsheet, sheet_name)))
        elif choice == '4':
            sheet_name = input("Enter the name of the sheet: ")
            values = get_sheet_values(spreadsheet, sheet_name)
            display_table(values)
        elif choice == '5':
            sheet_name = input("Enter the name of the sheet: ")
            analyze_data(spreadsheet, sheet_name)
        elif choice == '6':
            sheet_name = input("Enter the name of the sheet: ")
            export_to_xlsx(spreadsheet, sheet_name)
        elif choice == '7':
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please try again.")
    
    print_credit()

if __name__ == "__main__":
    main()
